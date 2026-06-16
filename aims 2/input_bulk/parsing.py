"""Parse + column-map bulk uploads (CSV / XLSX / JSONL) into canonical Story records.

Canonical columns: subject_id, picture, story_text — OR a StoryID like ``P010_PIC1`` from which
subject + picture are derived. Auto-detects common header names; the operator can override the
mapping before scoring.
"""
from __future__ import annotations

import csv
import io
import json
from typing import Optional

from core.ids import parse_story_id
from core.models import Story, count_words

# candidate header names (lowercased, non-alnum stripped) -> canonical field
_ALIASES = {
    "subject_id": {"subjectid", "subject", "subjid", "participant", "participantid", "pid", "person"},
    "picture": {"picture", "pic", "pictureindex", "picnum", "card", "image", "cue", "stimulus"},
    "story_text": {"storytext", "story", "text", "narrative", "response", "transcript", "content"},
    "story_id": {"storyid", "id", "story"},
}


def _norm(name: str) -> str:
    return "".join(ch for ch in (name or "").lower() if ch.isalnum())


def parse_upload(filename: str, content: bytes) -> tuple[list[dict], list[str]]:
    name = (filename or "").lower()
    if name.endswith(".jsonl") or name.endswith(".ndjson"):
        rows = [json.loads(line) for line in content.decode("utf-8", "replace").splitlines()
                if line.strip()]
        cols = list(rows[0].keys()) if rows else []
        return rows, cols
    if name.endswith(".xlsx"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header = [str(c) if c is not None else "" for c in next(it, [])]
        rows = [{header[i]: ("" if v is None else v) for i, v in enumerate(r) if i < len(header)}
                for r in it]
        return rows, header
    # default: CSV/TSV
    text = content.decode("utf-8", "replace")
    dialect = "excel-tab" if (name.endswith(".tsv") or "\t" in text[:1000]) else "excel"
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    rows = list(reader)
    return rows, list(reader.fieldnames or [])


def auto_map(columns: list[str]) -> dict[str, Optional[str]]:
    norm_to_orig = {_norm(c): c for c in columns}
    mapping: dict[str, Optional[str]] = {k: None for k in _ALIASES}
    for field, aliases in _ALIASES.items():
        for n, orig in norm_to_orig.items():
            if n in aliases:
                mapping[field] = orig
                break
    # if a StoryID column exists, prefer deriving subject/picture from it
    return mapping


def to_stories(rows: list[dict], mapping: dict[str, Optional[str]]) -> tuple[list[Story], list[str]]:
    errors: list[str] = []
    stories: list[Story] = []
    c_sid, c_pic, c_text, c_storyid = (mapping.get("subject_id"), mapping.get("picture"),
                                       mapping.get("story_text"), mapping.get("story_id"))
    for i, row in enumerate(rows, 1):
        text = str(row.get(c_text, "") if c_text else "").strip()
        subject = str(row.get(c_sid, "") if c_sid else "").strip()
        pic_raw = row.get(c_pic, "") if c_pic else ""
        story_id = str(row.get(c_storyid, "") if c_storyid else "").strip()

        if (not subject or not pic_raw) and story_id:
            pid = parse_story_id(story_id)
            subject = subject or pid.subject_id
            picture = int(pic_raw) if str(pic_raw).strip().isdigit() else pid.picture
        else:
            try:
                picture = int(float(pic_raw)) if str(pic_raw).strip() else 0
            except (TypeError, ValueError):
                picture = 0
        if not text:
            errors.append(f"Row {i}: empty story text — skipped.")
            continue
        if not subject:
            errors.append(f"Row {i}: no subject id (and no parseable StoryID) — skipped.")
            continue
        stories.append(Story.make(subject, picture, text, source="bulk"))
    return stories, errors


def parse_paste(blocks: str) -> tuple[list[Story], list[str]]:
    """Manual grouped paste. Format: lines ``subject_id | picture | story...`` OR blank-line-
    separated blocks beginning with ``# subject_id`` then one story per line."""
    stories, errors = [], []
    current = None
    pic = 0
    for ln, raw in enumerate(blocks.splitlines(), 1):
        line = raw.strip()
        if not line:
            continue
        if line.startswith("#"):
            current = line[1:].strip()
            pic = 0
            continue
        if "|" in line:
            parts = [p.strip() for p in line.split("|", 2)]
            if len(parts) == 3:
                subj, pictxt, text = parts
                try:
                    pnum = int(pictxt)
                except ValueError:
                    pnum = 0
                if text:
                    stories.append(Story.make(subj, pnum, text, source="bulk"))
                else:
                    errors.append(f"Line {ln}: empty story — skipped.")
                continue
        if current:
            pic += 1
            stories.append(Story.make(current, pic, line, source="bulk"))
        else:
            errors.append(f"Line {ln}: no subject context (use '# subject_id' header or "
                          f"'subject | pic | story') — skipped.")
    return stories, errors
